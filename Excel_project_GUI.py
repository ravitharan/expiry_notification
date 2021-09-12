from tkinter import *
import openpyxl  # A python module to handle Excel files
import datetime    # A python module to handle date and time
from tkinter import filedialog
def excel_file_handler():
    Path=lab.cget("text")
    Accno = path1.get()
    Valid = path2.get()
    Email = path3.get()
    #Accno = str(input("Enter the title of the area where the Access card numbers are : ")) 
    #Valid = str(input("Enter the title of the area where the Valid till date is : "))
    #Email = str(input("Enter the title of the area where the Email address is : "))
    user_excel_file = openpyxl.load_workbook(filename = Path)                                                       # load user excel file into memory
    sheet_names = user_excel_file.sheetnames   # create a list of sheet name's of user excel file
    for current_sheet in sheet_names:          # iterating sheet name of user excel file in current sheet
        active_sheet = user_excel_file[current_sheet]          # activate the current sheet for use it
        for row in active_sheet.rows:       # iterating whole excel data for check and get the cell adress of Access card no and Valid till
            for i in row:                   # iterating every row tuple
                if i.value == Accno:         # checking the value is Access card no ?
                    Access_card_no = str(i)             # convert the type <class 'openpyxl.cell.cell.Cell'> to <class 'str'>
                    get_address_of_cell = Access_card_no.split('.')   # split the string for filter cell adress 
                    length_of_list = len(get_address_of_cell)
                    Address = get_address_of_cell[length_of_list - 1].split('>') # split the string for filter cell adress
                    Address_of_Access_card_no = Address[0]          # now successfully got the adress of Access card no cell
                    data_of_Access_card_no = active_sheet[Address_of_Access_card_no[0]]       # storing the data of Access card no into a tuple
                    data_of_Access_card_no = list(data_of_Access_card_no)        # covert the tuple data of Access card no to a list for manipulate the data
                    
                elif i.value == Valid:
                    Valid_till = str(i)             # convert the type <class 'openpyxl.cell.cell.Cell'> to <class 'str'>
                    get_address_of_cell = Valid_till.split('.')   # split the string for filter cell adress 
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')   # split the string for filter cell adress 
                    Address_of_Valid_till = Address[0]           # now successfully got the adress of Valid_till cell
                    data_of_Valid_till = active_sheet[Address_of_Valid_till[0]]      # storing the data of Valid till into a tuple
                    data_of_Valid_till = list(data_of_Valid_till)           # convert the tuple data of Valid till to a list for manipulate the data
                elif i.value == Email:
                    Email_str = str(i)             # convert the type <class 'openpyxl.cell.cell.Cell'> to <class 'str'>
                    get_address_of_cell = Email_str.split('.')   # split the string for filter cell adress 
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')   # split the string for filter cell adress 
                    Address_of_Email_str = Address[0]           # now successfully got the adress of Email cell
                    data_of_Email_str = active_sheet[Address_of_Email_str[0]]      # storing the data of Email into a tuple
                    data_of_Email_str = list(data_of_Email_str)           # covert the tuple data of Email to a list for manipulate the data


    def merge(list1, list2):
                        merged_list = [(list1[i], list2[i]) for i in range(0, len(list1))]
                        return merged_list
    valid_till_and_email = merge(data_of_Valid_till, data_of_Email_str)
                   
    # Convert two lists into a dictionary
    dictionary = {}
    for key in data_of_Access_card_no:
        for value in valid_till_and_email:
            dictionary[key] = value
            valid_till_and_email.remove(value)
            break

    
    today = datetime.datetime.today()   # get current date
    today = today.replace(hour=0, minute=0, second=0, microsecond=0) # get date with zero time
    final_dictionary = {}           # create a dictionary to add name and remainig days

    # iterating dictionary through keys
    for key in dictionary:
        temp_value = dictionary[key][0].value      # Store the Valid till data into temp value
        if type(temp_value) == type(today):      # check the value type is date
            Valid_till_date = temp_value    # store the date value into Valid till date
            Remaining_days = Valid_till_date - today    # get the remaining days and store into Remaining days
            Remaining_days = Remaining_days.days        # convert <class 'datetime.timedelta'> to <class 'int'> and store into Remainig days
            if Remaining_days <= 5:                    # check the data is less than or equal to 5
                final_dictionary[key] = (Remaining_days,dictionary[key][1])    # if the value is less or equal to 5 then store into final dictionary
                        

                        
          
    # save the final dictionary as a excel file into Remainder.xlsx                              
    workbook = openpyxl.Workbook()             # open a Workbook as named work book
    sheet = workbook.active                    # active workbook's sheet for use it
    sheet["A1"] = "Access card no"             # save Access card number as a title into A1 cell 
    sheet["B1"] = "Remaining days"             # save Remaining days as a title into B1 cell
    sheet["C1"] = "Email"                      # save Email as a title into C1 cell
    i = 2                                      # increment value by 1 for access cell address 
    for key in final_dictionary:               # iterating final dictionary through key
        sheet["A" + str(i)] = key.value    # add dictionary key into Ax
        sheet["B" + str(i)] = final_dictionary[key][0]        # add dictionary value into Bx
        sheet["C" + str(i)] = final_dictionary[key][1].value       # add dictionary value into Cx
        i += 1                             # increment by 1
    workbook.save(filename='Remainder.xlsx')   # save the file into current directory



root = Tk()
root.title('File Explorer')
root.geometry("1400x1400")
root.config(background = "#158FAD")
path1 = StringVar()
path2 = StringVar()
path3 = StringVar()

def browseFiles():
    filename = filedialog.askopenfilename(initialdir = "/""/",title = "Select a File",filetypes = (("Text files","*.xlsx*"),("all files","*.*")))
    lab.configure(text=filename)
    
lab = Label(root,text="Select Your Excel File's Location",font=("Times New Roman",20,'bold'),height=3,bg='#158FAD')
lab.grid(row=0,column=2 )


button_explore = Button(root,font=("Elephant",20),text = "Browse Files",command =browseFiles)
button_explore.grid(row=1,column=2,sticky="ew")


lab1 =Label(root,text="Enter the title of the area where the Access card numbers are : ",font=("Elephant",20),bg="#158FAD",pady="10")
lab1.grid(row="3",column="1",sticky="ew")

txtName1 = Entry(root,font=("Elephant",20),textvariable = path1)
txtName1.grid(row="3",column="2",sticky="ew")

lab2 =Label(root,text="Enter the title of the area where the Valid till date is : ",font=("Elephant",20),bg="#158FAD",pady="10")
lab2.grid(row="4",column="1",sticky="ew")

txtName2 = Entry(root,font=("Elephant",20),textvariable = path2)
txtName2.grid(row="4",column="2",sticky="ew")

lab3 =Label(root,text="Enter the title of the area where the Email address is : ",font=("Elephant",20),bg="#158FAD",pady="10")
lab3.grid(row="5",column="1",sticky="ew")

txtName3 = Entry(root,font=("Elephant",20),textvariable = path3)
txtName3.grid(row="5",column="2",sticky="ew")

But = Button(root,text ="Run",fg="black", font=("Elephant",20),bg='#808080' , command = excel_file_handler)
But.grid(row="8",column="2",sticky="ew")

button_exit = Button(root,text = "Exit",font=("Elephant",20),command = exit,bg='#808080')
button_exit.grid(row="9",column="2",sticky="ew")

root.mainloop()






    


