import json
import openpyxl
import sys
from datetime import datetime

CONFIG_FILE_NAME = ".config.json"
OUTPUT_FILE_NAME = "Reminder.xlsx"

# Get following parameters from GUI inputs.
CONFIG_PARAMS = {
    "email" : "user@email.com",
    "reminder_days" : 5,
    "expiry_title" : "Valid till",
    "output_columns" : ["Name", "Access card no", "Vehicle no"],
}

def update_configfile(config_params, file_name):
    with open(file_name, 'w') as fp:
        json.dump(config_params, fp)

def retrive_configfile(file_name):
    with open(file_name) as fp:
        params = json.load(fp)
    return params

def parse_xlsx_header(xlsx_file, cfg_params):
    '''
    Parse xlsx file and returen column number for expiry date and other output
    columns. Also returning data start row number

    '''

    error_msg = None
    data_locations = {
        "work_sheet" : None,
        "data_start_row": None,
        "columns": None,
    }
    # First interested column is expiry date
    column_labels = [cfg_params["expiry_title"]]
    # Subsequent columns are requested by user
    column_labels.extend(cfg_params["output_columns"])

    data_locations["columns"] = [None for x in column_labels]

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    data_locations["work_sheet"] = ws

    for row in range(1, ws.max_row+1):
        for column in range(1, ws.max_column+1):
            for index, col_label in enumerate(column_labels):
                if ws.cell(row, column).value == col_label:
                    data_locations["columns"][index] = column
                    data_locations["data_start_row"] = row + 1
            if not None in data_locations["columns"]:
                break
        if not None in data_locations["columns"]:
            break
    if None in data_locations["columns"]: # One or more columns not found
        miss_columns = [ label for col, label in zip(data_locations["columns"], column_labels) if col == None]
        error_msg = "Cannot find columns: " + ", ".join(miss_columns)

    return (error_msg, data_locations)

def get_xlsx_data(xlsx_locations):
    '''
    Return expiry date and other output columns in a list from the excel file
    '''
    xlsx_data = []
    ws = xlsx_locations["work_sheet"]

    for row in range(xlsx_locations["data_start_row"], ws.max_row+1):
        row_data = []
        for column in xlsx_locations["columns"]:
            value = ws.cell(row, column).value
            row_data.append(value)
        if row_data[0] != None:
            xlsx_data.append(row_data)

    return xlsx_data

def filter_data(xlsx_data, num_expiry_days):
    '''
    Filter xlsx_data within num_expiry_days. Sort the filtered data 
    '''
    today = datetime.today()
    data = []
    for row_data in xlsx_data:
        remaining = row_data[0] - today
        if (remaining.days < num_expiry_days):
            row_data.insert(0, remaining.days)
            data.append(row_data)
    if data:
        return sorted(data, key=lambda x : x[0])
    else:
        return None

def write_xlsx_file(filtered_data, cfg_params):
    '''
    Write filtered data into output xlsx file
    '''
    output_file = OUTPUT_FILE_NAME
    wb = openpyxl.Workbook()             # open a Workbook as named work book
    ws = wb.active
    ws.cell(1, 1).value = "Remaining days"
    ws.cell(1, 2).value = cfg_params["expiry_title"]
    for col in range(3, len(cfg_params["output_columns"]) + 3):
        ws.cell(1, col).value = cfg_params["output_columns"][col-3]

    row = 1
    for row_data in filtered_data:
        row += 1
        for col_1, cell_data in enumerate(row_data):
            if col_1 == 1:
                ws.cell(row, col_1 + 1).value = cell_data.strftime("%Y-%b-%d")
            else:
                ws.cell(row, col_1 + 1).value = cell_data
    wb.save(output_file)   # save the file into current directory
    return output_file

def process_xlsx(input_xlsx, config):
    (err_msg, locations) = parse_xlsx_header(input_xlsx, config)
    if err_msg != None:
        return err_msg
    data = get_xlsx_data(locations)
    filtered_data = filter_data(data, config["reminder_days"])
    write_xlsx_file(filtered_data, config)
    return None

if __name__ == '__main__':

    if (len(sys.argv) != 2):
        print(f'Argument error\n Usage: {sys.argv[0]} <input_excel_file>')
        exit(1)
    err_msg = process_xlsx(sys.argv[1], CONFIG_PARAMS)
    if err_msg != None:
        print(err_msg)
